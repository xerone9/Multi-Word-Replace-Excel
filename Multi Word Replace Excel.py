from tkinter import *
import webbrowser
from tkinter import filedialog
import openpyxl as xl
from openpyxl.styles import Font
import os


def callback(url):
    webbrowser.open_new(url)


def uploadExcel():
    filetypes = (
        ('Excel 2010 Files', '*.xlsx'),
    )

    filename = filedialog.askopenfilename(
        title='Open a file',
        filetypes=filetypes)
    excelLocation.config(text=filename)
    getExcelLocation.config(foreground="black")
    startButton.place(x=165, y=545)
    errorText.place(x=1650, y=545)

def uploadDictionary():
    filetypes = (
        ('Excel 2010 Files', '*.xlsx'),
    )

    filename = filedialog.askopenfilename(
        title='Open a file',
        filetypes=filetypes)
    dictionaryLocation.config(text=filename)
    getDictionaryLocation.config(foreground="black")
    startButton.place(x=165, y=545)
    errorText.place(x=1650, y=545)


def multiWordReplaceExcel():
    locationExcel = excelLocation.cget("text")
    locationDictionary = dictionaryLocation.cget("text")
    pick = pickText.get()
    paste = pasteText.get()
    if len(locationExcel) < 1:
        getExcelLocation.config(foreground="red")

    elif len(locationDictionary) < 1:
        getDictionaryLocation.config(foreground="red")

    elif locationExcel == locationDictionary and len(locationExcel) > 0 and len(locationDictionary) > 0:
        startButton.place(x=1650, y=545)
        errorText.configure(text="Selected Files are same")
        errorText.place(x=105, y=545)
    elif len(pick) < 1 and len(paste) < 1:
        pickLabel.config(foreground="red")
        pasteLabel.config(foreground="red")
    elif len(pick) < 1:
        pickLabel.config(foreground="red")
        pasteLabel.config(foreground="black")
    elif len(paste) < 1:
        pickLabel.config(foreground="black")
        pasteLabel.config(foreground="red")
    else:
        # word = input("Enter Word: ");
        # result = NIC_Dictionary(word)
        # print(result)

        wb = xl.load_workbook(locationExcel, data_only=True)
        wc = xl.load_workbook(locationDictionary, data_only=True)
        #wb = xl.load_workbook(r'C:\Users\HYSTOU\Desktop\sest.xlsx')
        sheet = wb.worksheets[0]
        dictionarySheet = wc.worksheets[0]

        if pick != paste:
            for row in range(1, sheet.max_row+1):
                cell = sheet[str(pasteText.get())+str(row)]
                cell.value = 'Not Found'
                cell.font = Font(bold=True, name='Arial', color='FF0000')

        for row in range(1, sheet.max_row+1):
            cell = sheet[str(pickText.get())+str(row)]
            cellTranslatedInto = sheet[str(pasteText.get())+str(row)]
            for value in range(1, dictionarySheet.max_row+1):
                searchingWord = dictionarySheet.cell(value, 1)
                wordTranslation = dictionarySheet.cell(value, 2)


                # print(f"NIC File Value = {cell.value} and Dictionary File Value = {gel.value}")
                if cell.value == searchingWord.value:
                    cellTranslatedInto.value = wordTranslation.value
                    cellTranslatedInto.font = Font(bold=True, name='Arial', color='000000')
                    break
                else:
                    pass
                    # print("No Match")
            # NIC_Translation = NIC_Dictionary(cell.value)
            # NIC_Translated_cell = sheet.cell(row, 5)
            # NIC_Translated_cell.value = NIC_Translation

        while True:
            try:
                wb.save(locationExcel)
                os.startfile(locationExcel)
                print('done')
                break
            except PermissionError:
                print("")
                input("It seems the file is already open. Close the excel file and Press Enter to Conetinue...")

root = Tk()
root.resizable(0,0)
root.iconbitmap('icon.ico')
root.title('Amount Translator - V-2.0')
root.geometry("440x630")
root.configure(bg="white")

var = IntVar(None, 1)
location = ''


# myFont = font.Font(family='Helvetica', size=20, weight='bold')

img=PhotoImage(file='Logo-Microsoft-Excel.png')
label = Label(root, image=img)
label.configure(foreground="black")
label.configure(bg="white")
label.place(x=15, y=10)

getExcelLocation = Label(root, text="Select Excel File", font=("Sofia pro", 25, 'bold'))
getExcelLocation.configure(bg="white")
getExcelLocation.place(x=10, y=110)

browseButtonExcel = Button(root, text='Select File', font=("Sofia pro", 10, 'bold'), command=uploadExcel)
browseButtonExcel.configure(bg="light green")
browseButtonExcel.place(x=350, y=118)

excelLocation = Label(root, text="", font=("Roboto", 11))
excelLocation.config(bg='white')
excelLocation.config(foreground='dark green')
excelLocation.place(x=10, y=155)

getDictionaryLocation = Label(root, text="Select Dictionary", font=("Sofia pro", 25, 'bold'), justify='center')
getDictionaryLocation.configure(bg="white")
getDictionaryLocation.place(x=10, y=180)

browseButtonDictionary = Button(root, text='Select File', font=("Sofia pro", 10, 'bold'), command=uploadDictionary)
browseButtonDictionary.configure(bg="light green")
browseButtonDictionary.place(x=350, y=192)

dictionaryLocation = Label(root, text="", font=("Roboto", 11))
dictionaryLocation.config(bg='white')
dictionaryLocation.config(foreground='dark green')
dictionaryLocation.place(x=10, y=224)

textLabel = Label(root, text="Dictionary File:\nMake Sure that values needed to be translated are in first column of first sheet. Translated Values will be on second column Right Beside Them. THAT ARRANGEMENT IS MANDATORY", font=('Lato sans', 12), wraplength=420, justify='left', borderwidth=1, relief="solid", padx=7, pady=7)
textLabel.configure(bg="light grey")
textLabel.configure(bd=2)
textLabel.configure(foreground="green")
textLabel.place(x=10, y=260)


pickLabel = Label(root, text="Pick Values From Column: ", font=("Inter", 12, 'bold'), justify='center')
pickLabel.configure(bg="white")
pickLabel.place(x=10, y=373)

pickText = Entry(root, width=5, textvariable=(StringVar(root, value='A')), foreground='green', font=("Arial", 12, 'bold'))
pickText.place(x=230, y=374)

pasteLabel = Label(root, text="Paste Values Into Column: ", font=("Inter", 12, 'bold'), justify='center')
pasteLabel.configure(bg="white")
pasteLabel.place(x=10, y=407)

pasteText = Entry(root, width=5, textvariable=(StringVar(root, value='B')), foreground='green', font=("Arial", 12, 'bold'))
pasteText.place(x=230, y=408)

unitDescription = Label(root, text="Dictionary creation is fixed but the file you uploaded for translation has a flexibility that you can enter from which column the values should be picked and which column you want to place translated values. Data must be in First Sheet", font=(6), wraplength=420, justify='left', borderwidth=1, relief="solid", padx=7, pady=7)
unitDescription.configure(bg="light grey")
unitDescription.configure(bd=2)
unitDescription.configure(foreground="green")
unitDescription.place(x=10, y=440)

startButton = Button(root, text="S T A R T", font=("Arial", 15, 'bold'), justify='center', command=multiWordReplaceExcel)
startButton.configure(foreground="black")
startButton.configure(bg="light green")
startButton.place(x=165, y=545)

errorText = Label(root, text="", font=("Arial", 15, 'bold'), justify='center')
errorText.configure(foreground="red")
errorText.configure(bg="white")
errorText.place(x=1650, y=5450)

footer = Label(root, text="softwares.rubick.org", font=(14), cursor="hand2")
footer.bind("<Button-1>", lambda e: callback("http://softwares.rubick.org"))
footer.configure(foreground="white")
footer.configure(bg="black")
footer.pack(side=BOTTOM)
root.mainloop()

